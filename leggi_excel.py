import sys
import json
import openpyxl
from datetime import datetime

def leggi_classifica(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    squadre = []
    giornata = None

    for row in ws.iter_rows(values_only=True):
        if row[0] is None: continue
        if str(row[0]).strip() == 'Pos': continue
        try:
            pos = int(row[0])
            # colonna 1 = squadra, colonna 2 = vuota, colonna 3 = G
            squadra = str(row[1]).strip() if row[1] else None
            if not squadra: continue
            g   = int(row[3])  if row[3]  is not None else 0
            v   = int(row[4])  if row[4]  is not None else 0
            n   = int(row[5])  if row[5]  is not None else 0
            p   = int(row[6])  if row[6]  is not None else 0
            gf  = int(row[7])  if row[7]  is not None else 0
            gs  = int(row[8])  if row[8]  is not None else 0
            dr  = int(row[9])  if row[9]  is not None else 0
            pt  = int(row[10]) if row[10] is not None else 0
            ptt = float(row[11]) if row[11] is not None else 0.0
            squadre.append({
                "pos": pos, "squadra": squadra,
                "g": g, "v": v, "n": n, "p": p,
                "gf": gf, "gs": gs, "dr": dr,
                "punti": pt, "pt_totali": ptt
            })
            if g > 0 and giornata is None:
                giornata = g
        except (ValueError, TypeError):
            continue
    wb.close()
    return squadre, giornata

def leggi_calendario(path, giornata_num):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    risultati = []
    rows = list(ws.iter_rows(values_only=True))

    targets = [f"{giornata_num}ª Giornata lega", f"{giornata_num}a Giornata lega", f"{giornata_num}° Giornata lega"]

    for i, row in enumerate(rows):
        col_offset = None
        for col_idx, cell in enumerate(row):
            if cell and any(str(cell).startswith(t) for t in targets):
                col_offset = col_idx  # 0 = colonna sinistra, 6 = colonna destra
                break
        if col_offset is not None:
            for j in range(1, 5):
                if i + j >= len(rows): break
                r = rows[i + j]
                try:
                    casa  = str(r[col_offset]).strip() if r[col_offset] else None
                    pt_c  = float(r[col_offset+1]) if r[col_offset+1] is not None else 0
                    pt_f  = float(r[col_offset+2]) if r[col_offset+2] is not None else 0
                    fuori = str(r[col_offset+3]).strip() if r[col_offset+3] else None
                    ris   = str(r[col_offset+4]).strip() if r[col_offset+4] is not None else '-'
                    if not casa or not fuori or ris in ('-', 'None', ''): continue
                    gol = ris.split('-')
                    if len(gol) != 2: continue
                    risultati.append({
                        "casa": casa, "fuori": fuori,
                        "gol_casa": int(gol[0]), "gol_fuori": int(gol[1]),
                        "pt_casa": pt_c, "pt_fuori": pt_f
                    })
                except (ValueError, TypeError, IndexError):
                    continue
            break
    wb.close()
    return risultati

if __name__ == '__main__':
    classifica_path = sys.argv[1]
    calendario_path = sys.argv[2]
    output_path     = sys.argv[3]

    squadre, giornata = leggi_classifica(classifica_path)
    if not squadre:
        print("ERRORE: nessuna squadra trovata nella classifica")
        sys.exit(1)

    print(f"Squadre trovate: {len(squadre)}, giornata: {giornata}")

    risultati = leggi_calendario(calendario_path, giornata) if giornata else []
    print(f"Risultati ultima giornata: {len(risultati)}")

    fanta = {
        "giornata": f"Giornata {giornata}" if giornata else "In corso",
        "aggiornato_il": datetime.now().strftime('%Y-%m-%d'),
        "squadre": squadre,
        "ultima_giornata": {"numero": giornata, "risultati": risultati}
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(fanta, f, ensure_ascii=False, indent=2)

    print(f"OK: fantacalcio.json salvato")
